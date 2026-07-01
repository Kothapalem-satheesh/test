import io
from openpyxl import load_workbook
from app.services.chunker import chunk_text


def parse_excel(file_bytes: bytes) -> tuple[list[dict], int]:
    """Extract text from Excel spreadsheets."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    chunks = []
    total_sheets = len(wb.sheetnames)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_text = []

        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            if any(v.strip() for v in row_values):
                rows_text.append(" | ".join(row_values))

        if rows_text:
            sheet_text = f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text)
            sheet_chunks = chunk_text(
                sheet_text,
                section_title=f"Sheet: {sheet_name}",
            )
            chunks.extend(sheet_chunks)

    wb.close()

    for i, chunk in enumerate(chunks):
        chunk["chunk_index"] = i

    return chunks, total_sheets
